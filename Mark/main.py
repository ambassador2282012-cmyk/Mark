import os
import json
import customtkinter as ctk
from tkinter import ttk, messagebox, StringVar, IntVar, filedialog, simpledialog
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from docx import Document
except ImportError:
    Document = None

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

MONTHS = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]

EXPENSE_TYPES = ["Ремонт", "Электроэнергия", "Непредвиденные"]
STATE_FILE = "rent_app_state.json"


def money(value):
    return f"{value:.2f}"


class RentApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Mark")
        self.geometry("1500x920")
        self.minsize(1200, 780)
        self.resizable(True, True)

        self.rooms = []
        self.tenants = []
        self.room_vars = []
        self.records = []
        self.current_month = StringVar(value=MONTHS[0])

        self.selected_record_index = None
        self.detail_window = None
        self.edit_window = None

        self.bg_main = "#111318"
        self.bg_panel = "#181c23"
        self.bg_card = "#202632"
        self.bg_entry = "#242b38"
        self.accent = "#5f7ea8"
        self.accent_hover = "#4f6b90"
        self.success = "#5f8f6b"
        self.danger = "#8f6464"
        self.text = "#e6e8ec"
        self.muted = "#a4adbb"
        self.line = "#2d3442"

        self.configure(fg_color=self.bg_main)
        self.setup_tree_style()
        self.build_ui()
        self.load_state()
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def setup_tree_style(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure(
            "Treeview",
            background=self.bg_card,
            foreground=self.text,
            fieldbackground=self.bg_card,
            rowheight=33,
            borderwidth=0,
            relief="flat",
            font=("Segoe UI", 10)
        )

        style.configure(
            "Treeview.Heading",
            background=self.bg_panel,
            foreground=self.text,
            relief="flat",
            font=("Segoe UI", 10, "bold")
        )

        style.map(
            "Treeview",
            background=[("selected", self.accent)],
            foreground=[("selected", "white")]
        )

        style.map(
            "Treeview.Heading",
            background=[("active", self.accent_hover), ("!active", self.bg_panel)],
            foreground=[("active", "white"), ("!active", self.text)]
        )

    def build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.top_frame = ctk.CTkFrame(self, fg_color=self.bg_panel, corner_radius=16)
        self.top_frame.grid(row=0, column=0, padx=16, pady=(16, 10), sticky="ew")
        self.top_frame.grid_columnconfigure(1, weight=1)

        self.title_label = ctk.CTkLabel(
            self.top_frame,
            text="🏢 Mark — учёт аренды помещений",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color=self.text
        )
        self.title_label.grid(row=0, column=0, padx=18, pady=18, sticky="w")

        self.new_file_button = ctk.CTkButton(
            self.top_frame, text="🗂 Создать файл", command=self.create_new_file,
            fg_color=self.bg_entry, hover_color="#2d3646", text_color=self.text,
            corner_radius=12, height=38
        )
        self.new_file_button.grid(row=0, column=1, padx=8, pady=18, sticky="e")

        self.import_excel_button = ctk.CTkButton(
            self.top_frame, text="📥 Импорт Excel", command=self.import_excel,
            fg_color=self.bg_entry, hover_color="#2d3646", text_color=self.text,
            corner_radius=12, height=38
        )
        self.import_excel_button.grid(row=0, column=2, padx=8, pady=18, sticky="e")

        self.import_word_button = ctk.CTkButton(
            self.top_frame, text="📄 Импорт Word", command=self.import_word,
            fg_color=self.bg_entry, hover_color="#2d3646", text_color=self.text,
            corner_radius=12, height=38
        )
        self.import_word_button.grid(row=0, column=3, padx=8, pady=18, sticky="e")

        self.save_excel_button = ctk.CTkButton(
            self.top_frame, text="💾 Сохранить в Excel", command=self.save_to_excel,
            fg_color=self.bg_entry, hover_color="#2d3646", text_color=self.text,
            corner_radius=12, height=38
        )
        self.save_excel_button.grid(row=0, column=4, padx=(8, 18), pady=18, sticky="e")

        self.main_frame = ctk.CTkFrame(self, fg_color=self.bg_main)
        self.main_frame.grid(row=1, column=0, padx=16, pady=(0, 16), sticky="nsew")
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(2, weight=1)

        self.build_input_section()
        self.build_actions_section()
        self.build_table_section()
        self.build_summary_section()

    def build_input_section(self):
        self.input_frame = ctk.CTkFrame(self.main_frame, fg_color=self.bg_panel, corner_radius=16)
        self.input_frame.grid(row=0, column=0, padx=0, pady=(0, 12), sticky="ew")
        self.input_frame.grid_columnconfigure(1, weight=1)
        self.input_frame.grid_columnconfigure(3, weight=1)
        self.input_frame.grid_columnconfigure(5, weight=1)

        self.form_title = ctk.CTkLabel(
            self.input_frame,
            text="⚙ Сначала создайте файл и укажите помещения и арендаторов",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=self.text
        )
        self.form_title.grid(row=0, column=0, columnspan=6, padx=16, pady=(14, 10), sticky="w")

        self.rooms_value_label = ctk.CTkLabel(self.input_frame, text="Помещения: не заданы", text_color=self.muted)
        self.rooms_value_label.grid(row=1, column=0, columnspan=6, padx=16, pady=4, sticky="w")

        self.tenants_value_label = ctk.CTkLabel(self.input_frame, text="Арендаторы: не заданы", text_color=self.muted)
        self.tenants_value_label.grid(row=2, column=0, columnspan=6, padx=16, pady=(0, 10), sticky="w")

        ctk.CTkLabel(self.input_frame, text="📅 Месяц:", text_color=self.text).grid(row=3, column=0, padx=16, pady=8, sticky="w")
        self.month_menu = ctk.CTkOptionMenu(
            self.input_frame, values=MONTHS, variable=self.current_month,
            fg_color=self.bg_entry, button_color=self.accent, button_hover_color=self.accent_hover
        )
        self.month_menu.grid(row=3, column=1, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(self.input_frame, text="👤 Арендатор:", text_color=self.text).grid(row=3, column=2, padx=16, pady=8, sticky="w")
        self.tenant_menu = ctk.CTkOptionMenu(
            self.input_frame, values=["-"],
            fg_color=self.bg_entry, button_color=self.accent, button_hover_color=self.accent_hover
        )
        self.tenant_menu.grid(row=3, column=3, padx=10, pady=8, sticky="ew")

        self.add_tenant_to_list_button = ctk.CTkButton(
            self.input_frame,
            text="➕ Добавить арендатора",
            command=self.add_tenant_to_list,
            fg_color=self.bg_entry,
            hover_color="#2d3646",
            text_color=self.text,
            corner_radius=12,
            height=38
        )
        self.add_tenant_to_list_button.grid(row=3, column=4, columnspan=2, padx=16, pady=8, sticky="ew")

        ctk.CTkLabel(self.input_frame, text="💰 Сумма аренды:", text_color=self.text).grid(row=4, column=0, padx=16, pady=8, sticky="w")
        self.rent_entry = ctk.CTkEntry(
            self.input_frame,
            placeholder_text="например 30000.50",
            fg_color=self.bg_entry,
            border_color=self.line,
            text_color=self.text
        )
        self.rent_entry.grid(row=4, column=1, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(self.input_frame, text="📦 Помещения арендатора:", text_color=self.text).grid(row=4, column=2, padx=16, pady=8, sticky="w")

        self.rooms_check_frame = ctk.CTkFrame(self.input_frame, fg_color=self.bg_panel)
        self.rooms_check_frame.grid(row=5, column=0, columnspan=6, padx=16, pady=(0, 10), sticky="ew")

        self.add_button = ctk.CTkButton(
            self.input_frame,
            text="➕ Добавить запись аренды",
            command=self.add_tenant,
            fg_color=self.accent,
            hover_color=self.accent_hover,
            text_color="white",
            corner_radius=12,
            height=40
        )
        self.add_button.grid(row=6, column=0, columnspan=6, padx=16, pady=(8, 14), sticky="ew")

    def build_actions_section(self):
        self.actions_frame = ctk.CTkFrame(self.main_frame, fg_color=self.bg_panel, corner_radius=16)
        self.actions_frame.grid(row=1, column=0, padx=0, pady=(0, 12), sticky="ew")
        self.actions_frame.grid_columnconfigure((0, 1, 2), weight=1)

        self.show_button = ctk.CTkButton(self.actions_frame, text="🔎 Показать", command=self.show_record_details, fg_color=self.bg_entry, hover_color="#2d3646", text_color=self.text, corner_radius=12, height=38)
        self.show_button.grid(row=0, column=0, padx=12, pady=12, sticky="ew")

        self.edit_button = ctk.CTkButton(self.actions_frame, text="✏ Редактировать", command=self.edit_selected_record, fg_color=self.bg_entry, hover_color="#2d3646", text_color=self.text, corner_radius=12, height=38)
        self.edit_button.grid(row=0, column=1, padx=12, pady=12, sticky="ew")

        self.delete_button = ctk.CTkButton(self.actions_frame, text="🗑 Удалить", command=self.delete_selected_record, fg_color=self.bg_entry, hover_color="#3b2e34", text_color=self.text, corner_radius=12, height=38)
        self.delete_button.grid(row=0, column=2, padx=12, pady=12, sticky="ew")

    def build_table_section(self):
        self.table_frame = ctk.CTkFrame(self.main_frame, fg_color=self.bg_panel, corner_radius=16)
        self.table_frame.grid(row=2, column=0, padx=0, pady=(0, 12), sticky="nsew")
        self.table_frame.grid_columnconfigure(0, weight=1)
        self.table_frame.grid_rowconfigure(0, weight=1)

        columns = ("month", "tenant", "rooms", "rent", "expenses", "net")
        self.table = ttk.Treeview(self.table_frame, columns=columns, show="headings", height=14)

        headings = {
            "month": "Месяц",
            "tenant": "Арендатор",
            "rooms": "Помещения",
            "rent": "Аренда",
            "expenses": "Всего расходов",
            "net": "Чистый доход",
        }

        widths = {"month": 120, "tenant": 220, "rooms": 390, "rent": 130, "expenses": 130, "net": 130}

        for col in columns:
            self.table.heading(col, text=headings[col])
            self.table.column(col, anchor="center", width=widths[col], stretch=False)

        self.table.bind("<<TreeviewSelect>>", self.on_table_select)
        self.table.bind("<Double-1>", lambda e: self.show_record_details())

        vsb = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.table.yview)
        hsb = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.table.xview)
        self.table.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.table.grid(row=0, column=0, sticky="nsew", padx=12, pady=(12, 0))
        vsb.grid(row=0, column=1, sticky="ns", pady=(12, 0))
        hsb.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 12))

    def build_summary_section(self):
        self.summary_frame = ctk.CTkFrame(self.main_frame, fg_color=self.bg_panel, corner_radius=16)
        self.summary_frame.grid(row=3, column=0, padx=0, pady=(0, 0), sticky="ew")
        self.summary_frame.grid_columnconfigure(0, weight=1)

        self.summary_title = ctk.CTkLabel(self.summary_frame, text="📊 Итоги", font=ctk.CTkFont(size=16, weight="bold"), text_color=self.text)
        self.summary_title.grid(row=0, column=0, padx=16, pady=(12, 6), sticky="w")

        self.summary_label = ctk.CTkLabel(self.summary_frame, text="Данные пока не рассчитаны", justify="left", text_color=self.muted)
        self.summary_label.grid(row=1, column=0, padx=16, pady=(0, 14), sticky="w")

    def rebuild_room_checkboxes(self):
        for widget in self.rooms_check_frame.winfo_children():
            widget.destroy()
        self.room_vars = []
        for i, room in enumerate(self.rooms):
            var = IntVar(value=0)
            cb = ctk.CTkCheckBox(self.rooms_check_frame, text=room, variable=var)
            cb.grid(row=i // 4, column=i % 4, padx=12, pady=8, sticky="w")
            self.room_vars.append(var)

    def create_new_file(self):
        count_text = simpledialog.askstring("Новый файл", "Сколько помещений в объекте?")
        if not count_text:
            return
        try:
            count = int(count_text)
            if count <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректное число помещений.")
            return

        rooms = []
        for i in range(count):
            room_name = simpledialog.askstring("Названия помещений", f"Введите название помещения №{i + 1}:")
            if room_name is None or not room_name.strip():
                messagebox.showwarning("Внимание", "Название помещения не введено.")
                return
            rooms.append(room_name.strip())

        tenants_text = simpledialog.askstring("Арендаторы", "Введите имена арендаторов через запятую:")
        tenants = [t.strip() for t in tenants_text.split(",") if t.strip()] if tenants_text else []

        self.rooms = rooms
        self.tenants = tenants
        self.rooms_value_label.configure(text="Помещения: " + ", ".join(self.rooms))
        self.update_tenant_menu()
        self.form_title.configure(text="Файл создан. Можно добавлять арендаторов и расходы.")
        self.rebuild_room_checkboxes()
        self.refresh_table()

    def update_tenant_menu(self):
        self.tenant_menu.configure(values=self.tenants if self.tenants else ["-"])
        self.tenant_menu.set(self.tenants[0] if self.tenants else "-")
        self.tenants_value_label.configure(text="Арендаторы: " + ", ".join(self.tenants) if self.tenants else "Арендаторы: не заданы")

    def add_tenant_to_list(self):
        name = simpledialog.askstring("Новый арендатор", "Введите имя арендатора:")
        if not name or not name.strip():
            return
        name = name.strip()
        if name not in self.tenants:
            self.tenants.append(name)
            self.update_tenant_menu()
        self.tenant_menu.set(name)

    def add_tenant(self):
        if not self.rooms:
            messagebox.showwarning("Внимание", "Сначала создайте файл и укажите помещения.")
            return

        tenant = self.tenant_menu.get().strip()
        rent_text = self.rent_entry.get().strip()
        selected_rooms = [room for room, var in zip(self.rooms, self.room_vars) if var.get() == 1]

        if not tenant or tenant == "-" or not rent_text or not selected_rooms:
            messagebox.showerror("Ошибка", "Выберите арендатора, сумму и помещения.")
            return

        try:
            rent = float(rent_text.replace(",", "."))
        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректную сумму аренды, например 30000.50.")
            return

        record = {
            "month": self.current_month.get(),
            "tenant": tenant,
            "rooms": selected_rooms,
            "rent": rent,
            "share": rent / len(selected_rooms),
            "expenses": defaultdict(lambda: {t: 0.0 for t in EXPENSE_TYPES}),
        }
        self.records.append(record)
        self.refresh_table()
        self.clear_tenant_form()

    def clear_tenant_form(self):
        self.rent_entry.delete(0, "end")
        for var in self.room_vars:
            var.set(0)

    def calc_record(self, record):
        repair_total = sum(record["expenses"][room]["Ремонт"] for room in record["rooms"])
        electricity_total = sum(record["expenses"][room]["Электроэнергия"] for room in record["rooms"])
        unexpected_total = sum(record["expenses"][room]["Непредвиденные"] for room in record["rooms"])
        expenses = repair_total + electricity_total + unexpected_total
        net = record["rent"] - expenses
        return repair_total, electricity_total, unexpected_total, expenses, net

    def refresh_table(self):
        for item in self.table.get_children():
            self.table.delete(item)

        total_rent = 0.0
        total_expenses = 0.0

        for idx, record in enumerate(self.records):
            _, _, _, expenses, net = self.calc_record(record)
            total_rent += record["rent"]
            total_expenses += expenses

            self.table.insert("", "end", values=(
                record["month"], record["tenant"], ", ".join(record["rooms"]),
                money(record["rent"]), money(expenses), money(net)
            ), tags=(str(idx),))

        total_rooms = len(self.rooms)
        avg_without = total_rent / total_rooms if total_rooms else 0
        avg_with = (total_rent - total_expenses) / total_rooms if total_rooms else 0

        self.summary_label.configure(text=(
            f"Общий доход без расходов: {money(total_rent)}\n"
            f"Общий доход с расходами: {money(total_rent - total_expenses)}\n"
            f"Средний доход без расходов на помещение: {money(avg_without)}\n"
            f"Средний доход с расходами на помещение: {money(avg_with)}\n"
            f"Общее количество помещений: {total_rooms}"
        ), justify="left")

    def on_table_select(self, event=None):
        sel = self.table.selection()
        if not sel:
            self.selected_record_index = None
            return
        try:
            self.selected_record_index = int(self.table.item(sel[0], "tags")[0])
        except Exception:
            self.selected_record_index = None

    def selected_record(self):
        if self.selected_record_index is None or self.selected_record_index >= len(self.records):
            return None
        return self.records[self.selected_record_index]

    def show_record_details(self):
        record = self.selected_record()
        if record is None:
            messagebox.showwarning("Внимание", "Выберите запись в таблице.")
            return
        self.open_detail_window(record)

    def open_detail_window(self, record):
        if self.detail_window is not None and self.detail_window.winfo_exists():
            self.detail_window.destroy()

        self.detail_window = ctk.CTkToplevel(self)
        self.detail_window.title("Детали записи")
        self.detail_window.geometry("760x650")
        self.detail_window.resizable(True, True)
        self.detail_window.transient(self)
        self.detail_window.grab_set()
        self.detail_window.configure(fg_color=self.bg_main)

        frame = ctk.CTkScrollableFrame(self.detail_window, fg_color=self.bg_panel, corner_radius=16)
        frame.pack(fill="both", expand=True, padx=16, pady=16)

        repair_total, electricity_total, unexpected_total, expenses, net = self.calc_record(record)

        for text in [
            f"Месяц: {record['month']}",
            f"Арендатор: {record['tenant']}",
            f"Помещения: {', '.join(record['rooms'])}",
            f"Аренда: {money(record['rent'])}",
            f"Доля за помещение: {money(record['share'])}",
        ]:
            ctk.CTkLabel(frame, text=text, anchor="w", justify="left", text_color=self.text).pack(fill="x", pady=3)

        ctk.CTkLabel(frame, text="Расходы по помещениям:", font=ctk.CTkFont(weight="bold"), text_color=self.text).pack(anchor="w", pady=(14, 6))

        for room in record["rooms"]:
            r = record["expenses"][room]["Ремонт"]
            e = record["expenses"][room]["Электроэнергия"]
            u = record["expenses"][room]["Непредвиденные"]
            ctk.CTkLabel(frame, text=f"{room} | Ремонт: {money(r)} | Электроэнергия: {money(e)} | Непредвиденные: {money(u)}", anchor="w", justify="left", wraplength=700, text_color=self.text).pack(fill="x", pady=2)

        ctk.CTkLabel(frame, text=f"Ремонт всего: {money(repair_total)}", anchor="w", text_color=self.text).pack(fill="x", pady=(14, 2))
        ctk.CTkLabel(frame, text=f"Электроэнергия всего: {money(electricity_total)}", anchor="w", text_color=self.text).pack(fill="x", pady=2)
        ctk.CTkLabel(frame, text=f"Непредвиденные всего: {money(unexpected_total)}", anchor="w", text_color=self.text).pack(fill="x", pady=2)
        ctk.CTkLabel(frame, text=f"Всего расходов: {money(expenses)}", anchor="w", text_color=self.text).pack(fill="x", pady=2)
        ctk.CTkLabel(frame, text=f"Чистый доход: {money(net)}", anchor="w", font=ctk.CTkFont(weight="bold"), text_color=self.success).pack(fill="x", pady=(8, 2))

        ctk.CTkButton(frame, text="✖ Закрыть", command=self.detail_window.destroy, fg_color=self.bg_entry, hover_color="#2d3646", text_color=self.text, corner_radius=12).pack(pady=18)

    def edit_selected_record(self):
        record = self.selected_record()
        if record is None:
            messagebox.showwarning("Внимание", "Выберите запись в таблице.")
            return

        if self.edit_window is not None and self.edit_window.winfo_exists():
            self.edit_window.destroy()

        self.edit_window = ctk.CTkToplevel(self)
        self.edit_window.title("Редактирование записи")
        self.edit_window.geometry("960x820")
        self.edit_window.resizable(True, True)
        self.edit_window.transient(self)
        self.edit_window.grab_set()
        self.edit_window.configure(fg_color=self.bg_main)

        frame = ctk.CTkScrollableFrame(self.edit_window, fg_color=self.bg_panel, corner_radius=16)
        frame.pack(fill="both", expand=True, padx=16, pady=16)

        month_var = StringVar(value=record["month"])
        tenant_var = StringVar(value=record["tenant"])
        rent_var = StringVar(value=money(record["rent"]))

        ctk.CTkLabel(frame, text="📅 Месяц:", text_color=self.text).pack(anchor="w", pady=(8, 0))
        ctk.CTkOptionMenu(frame, values=MONTHS, variable=month_var, fg_color=self.bg_entry, button_color=self.accent, button_hover_color=self.accent_hover).pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(frame, text="👤 Арендатор:", text_color=self.text).pack(anchor="w", pady=(6, 0))
        ctk.CTkOptionMenu(frame, values=self.tenants if self.tenants else ["-"], variable=tenant_var, fg_color=self.bg_entry, button_color=self.accent, button_hover_color=self.accent_hover).pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(frame, text="💰 Сумма аренды:", text_color=self.text).pack(anchor="w", pady=(6, 0))
        ctk.CTkEntry(frame, textvariable=rent_var, fg_color=self.bg_entry, border_color=self.line, text_color=self.text).pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(frame, text="📦 Помещения:", font=ctk.CTkFont(weight="bold"), text_color=self.text).pack(anchor="w", pady=(12, 6))

        room_vars = {}
        for room in self.rooms:
            var = IntVar(value=1 if room in record["rooms"] else 0)
            ctk.CTkCheckBox(frame, text=room, variable=var).pack(anchor="w", pady=2)
            room_vars[room] = var

        ctk.CTkLabel(frame, text="Расходы по помещениям:", font=ctk.CTkFont(weight="bold"), text_color=self.text).pack(anchor="w", pady=(16, 6))

        expense_vars = {}
        for room in self.rooms:
            room_frame = ctk.CTkFrame(frame, fg_color=self.bg_card, corner_radius=12)
            room_frame.pack(fill="x", pady=6, padx=2)
            room_frame.grid_columnconfigure((1, 2, 3), weight=1)

            ctk.CTkLabel(room_frame, text=room, width=150, text_color=self.text).grid(row=0, column=0, padx=10, pady=(10, 4), sticky="w")

            labels_frame = ctk.CTkFrame(room_frame, fg_color="transparent")
            labels_frame.grid(row=1, column=0, columnspan=4, padx=10, pady=(0, 2), sticky="ew")
            labels_frame.grid_columnconfigure((0, 1, 2), weight=1)

            values_frame = ctk.CTkFrame(room_frame, fg_color="transparent")
            values_frame.grid(row=2, column=0, columnspan=4, padx=10, pady=(0, 10), sticky="ew")
            values_frame.grid_columnconfigure((0, 1, 2), weight=1)

            vars_for_room = {}
            for col, exp_type in enumerate(EXPENSE_TYPES):
                ctk.CTkLabel(labels_frame, text=exp_type, text_color=self.muted, anchor="center").grid(row=0, column=col, padx=6, sticky="ew")
                value = record["expenses"][room][exp_type] if room in record["rooms"] else 0.0
                var = StringVar(value=str(value))
                ctk.CTkEntry(values_frame, textvariable=var, fg_color=self.bg_entry, border_color=self.line, text_color=self.text).grid(row=0, column=col, padx=6, sticky="ew")
                vars_for_room[exp_type] = var

            expense_vars[room] = vars_for_room

        def save_changes():
            new_month = month_var.get().strip()
            new_tenant = tenant_var.get().strip()
            new_rent_text = rent_var.get().strip()
            new_rooms = [room for room, var in room_vars.items() if var.get() == 1]

            if not new_month or not new_tenant or not new_rooms:
                messagebox.showerror("Ошибка", "Заполните месяц, арендатора и выберите помещения.")
                return

            try:
                new_rent = float(new_rent_text.replace(",", "."))
            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректную сумму аренды с копейками.")
                return

            new_expenses = defaultdict(lambda: {t: 0.0 for t in EXPENSE_TYPES})
            for room in self.rooms:
                if room not in new_rooms:
                    continue
                for exp_type in EXPENSE_TYPES:
                    try:
                        new_expenses[room][exp_type] = float(expense_vars[room][exp_type].get().strip().replace(",", ".") or "0")
                    except ValueError:
                        messagebox.showerror("Ошибка", f"Некорректное значение: {room} — {exp_type}")
                        return

            record["month"] = new_month
            record["tenant"] = new_tenant
            record["rooms"] = new_rooms
            record["rent"] = new_rent
            record["share"] = new_rent / len(new_rooms)
            record["expenses"] = new_expenses

            self.refresh_table()
            self.edit_window.destroy()

        ctk.CTkButton(frame, text="💾 Сохранить изменения", command=save_changes, fg_color=self.bg_entry, hover_color="#2d3646", text_color=self.text, corner_radius=12).pack(pady=18)

    def delete_selected_record(self):
        idx = self.selected_record_index
        if idx is None or idx < 0 or idx >= len(self.records):
            messagebox.showwarning("Внимание", "Выберите запись в таблице.")
            return
        if not messagebox.askyesno("Подтверждение", "Удалить выбранную запись?"):
            return
        del self.records[idx]
        self.selected_record_index = None
        self.refresh_table()

    def import_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")], title="Выберите Excel файл")
        if not path:
            return
        try:
            wb = load_workbook(path)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось прочитать Excel: {e}")
            return

        self.records.clear()
        self.rooms = []
        self.tenants = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            if len(row) < 4:
                continue
            month = str(row[0]).strip()
            tenant = str(row[1]).strip()
            rooms = [x.strip() for x in str(row[2]).split(",") if x.strip()]
            try:
                rent = float(str(row[3]).replace(",", "."))
            except Exception:
                continue
            self.records.append({
                "month": month,
                "tenant": tenant,
                "rooms": rooms,
                "rent": rent,
                "share": rent / len(rooms) if rooms else rent,
                "expenses": defaultdict(lambda: {t: 0.0 for t in EXPENSE_TYPES}),
            })
            if tenant not in self.tenants:
                self.tenants.append(tenant)
            for room in rooms:
                if room not in self.rooms:
                    self.rooms.append(room)

        self.rooms_value_label.configure(text="Помещения: " + ", ".join(self.rooms) if self.rooms else "Помещения: не заданы")
        self.update_tenant_menu()
        self.rebuild_room_checkboxes()
        self.refresh_table()

    def import_word(self):
        if Document is None:
            messagebox.showerror("Ошибка", "Не установлен пакет python-docx.")
            return

        path = filedialog.askopenfilename(filetypes=[("Word files", "*.docx")], title="Выберите Word файл")
        if not path:
            return

        try:
            doc = Document(path)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось прочитать Word: {e}")
            return

        self.records.clear()
        self.rooms = []
        self.tenants = []

        for table in doc.tables:
            if not table.rows:
                continue
            for row in table.rows[1:]:
                data = [cell.text.strip() for cell in row.cells]
                if len(data) < 4:
                    continue
                month, tenant, rooms_text, rent_text = data[0], data[1], data[2], data[3]
                rooms = [x.strip() for x in rooms_text.split(",") if x.strip()]
                try:
                    rent = float(rent_text.replace(",", "."))
                except Exception:
                    continue

                self.records.append({
                    "month": month,
                    "tenant": tenant,
                    "rooms": rooms,
                    "rent": rent,
                    "share": rent / len(rooms) if rooms else rent,
                    "expenses": defaultdict(lambda: {t: 0.0 for t in EXPENSE_TYPES}),
                })

                if tenant not in self.tenants:
                    self.tenants.append(tenant)
                for room in rooms:
                    if room not in self.rooms:
                        self.rooms.append(room)

            if self.records:
                break

        self.rooms_value_label.configure(text="Помещения: " + ", ".join(self.rooms) if self.rooms else "Помещения: не заданы")
        self.update_tenant_menu()
        self.rebuild_room_checkboxes()
        self.refresh_table()

    def save_state(self):
        data = {
            "rooms": self.rooms,
            "tenants": self.tenants,
            "current_month": self.current_month.get(),
            "records": []
        }
        for rec in self.records:
            expenses = {room: dict(rec["expenses"][room]) for room in rec["rooms"]}
            data["records"].append({
                "month": rec["month"],
                "tenant": rec["tenant"],
                "rooms": rec["rooms"],
                "rent": rec["rent"],
                "share": rec["share"],
                "expenses": expenses
            })

        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def load_state(self):
        if not os.path.exists(STATE_FILE):
            self.refresh_table()
            return
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            self.refresh_table()
            return

        self.rooms = data.get("rooms", [])
        self.tenants = data.get("tenants", [])
        self.current_month.set(data.get("current_month", MONTHS[0]))
        self.records = []

        for rec in data.get("records", []):
            expenses = defaultdict(lambda: {t: 0.0 for t in EXPENSE_TYPES})
            for room, vals in rec.get("expenses", {}).items():
                expenses[room] = {
                    "Ремонт": float(vals.get("Ремонт", 0)),
                    "Электроэнергия": float(vals.get("Электроэнергия", 0)),
                    "Непредвиденные": float(vals.get("Непредвиденные", 0)),
                }
            self.records.append({
                "month": rec.get("month", MONTHS[0]),
                "tenant": rec.get("tenant", ""),
                "rooms": rec.get("rooms", []),
                "rent": float(rec.get("rent", 0)),
                "share": float(rec.get("share", 0)),
                "expenses": expenses
            })

        self.rooms_value_label.configure(text="Помещения: " + ", ".join(self.rooms) if self.rooms else "Помещения: не заданы")
        self.update_tenant_menu()
        self.rebuild_room_checkboxes()
        self.refresh_table()

    def on_close(self):
        try:
            self.save_state()
        except Exception:
            pass
        self.destroy()

    def save_to_excel(self):
        if not self.records:
            messagebox.showwarning("Внимание", "Нет данных для сохранения.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Сохранить как")
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Аренда"

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["Месяц", "Арендатор", "Помещения", "Аренда", "Всего расходов", "Чистый доход"]

        for col, head in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=head)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        row = 2
        total_rent = 0.0
        total_expenses = 0.0

        for rec in self.records:
            _, _, _, expenses, net = self.calc_record(rec)
            total_rent += rec["rent"]
            total_expenses += expenses

            values = [rec["month"], rec["tenant"], ", ".join(rec["rooms"]), rec["rent"], expenses, net]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = left if col in (2, 3) else center
                if col in (4, 5, 6):
                    cell.number_format = '#,##0.00'
            row += 1

        total_rooms = len(self.rooms)
        avg_without = total_rent / total_rooms if total_rooms else 0
        avg_with = (total_rent - total_expenses) / total_rooms if total_rooms else 0

        row += 1
        ws.cell(row=row, column=1, value="Итоги").font = Font(bold=True, size=14)
        row += 1

        for name, value in [
            ("Общий доход без расходов", total_rent),
            ("Общий доход с расходами", total_rent - total_expenses),
            ("Средний доход без расходов на помещение", avg_without),
            ("Средний доход с расходами на помещение", avg_with),
            ("Общее количество помещений", total_rooms),
        ]:
            ws.cell(row=row, column=1, value=name).font = bold_font
            ws.cell(row=row, column=2, value=value).number_format = '#,##0.00'
            row += 1

        widths = {1: 18, 2: 28, 3: 35, 4: 15, 5: 18, 6: 18}
        for col_idx, width in widths.items():
            ws.column_dimensions[chr(64 + col_idx)].width = width

        wb.save(file_path)
        messagebox.showinfo("Готово", f"Файл сохранён:\n{file_path}")


if __name__ == "__main__":
    app = RentApp()
    app.mainloop()
